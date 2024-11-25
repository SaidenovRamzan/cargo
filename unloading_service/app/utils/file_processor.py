import openpyxl
from io import BytesIO
from typing import List


def parse_excel_file(file_contents: bytes) -> List[dict]:
    """
    Парсит Excel файл и возвращает список данных из строк.

    :param file_contents: Содержимое файла в виде байтов.
    :return: Список словарей с данными из файла.
    """
    try:
        workbook = openpyxl.load_workbook(BytesIO(file_contents))
        sheet = workbook.active

        # Пропускаем строку с заголовками и собираем данные
        data = []
        for row in sheet.iter_rows(min_row=3, values_only=True):
            if not row[1]:  # Пропускаем строки с пустым vehicle_number
                continue

            data.append({
                "tracking_number": row[3],
                "quantity": int(row[4]),
            })
        return data
    except Exception as e:
        raise ValueError(f"Ошибка обработки файла: {e}")


def parse_excel_cargo_file(file_contents: bytes) -> List[dict]:
    """
    Парсит Excel файл и возвращает список данных из строк.

    :param file_contents: Содержимое файла в виде байтов.
    :return: Список словарей с данными из файла.
    """
    try:
        workbook = openpyxl.load_workbook(BytesIO(file_contents))
        sheet = workbook.active

        # Пропускаем строку с заголовками и собираем данные
        data = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row[1]:  # Пропускаем строки с пустым vehicle_number
                continue
            print('='*100)
            print(row[2])
            print(row[1])
            print('='*100)


            data.append({
                "tracking_number": row[1],
                "quantity": int(row[2]),
            })
        return data
    except Exception as e:
        raise ValueError(f"Ошибка обработки файла: {e}")
